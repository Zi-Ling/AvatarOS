# app/avatar/skills/builtin/excel.py

from __future__ import annotations

import logging
from typing import Any, List, Optional, Tuple, Union
import json
from pathlib import Path
from pydantic import Field, model_validator
from openpyxl import load_workbook, Workbook

from ..common.path_normalizer import normalize_file_extension
from ..common.path_mixins import PathBindMixin
from ..base import (
    BaseSkill,
    SkillSpec,
    SkillCategory,
    SkillPermission,
    SkillMetadata,
    SkillDomain,
    SkillCapability,
)

logger = logging.getLogger(__name__)
from ..schema import SkillInput, SkillOutput
from ..registry import register_skill
from ..context import SkillContext

import os


# =============================================================================
# Helpers
# =============================================================================

def _load_or_create_workbook(path: Union[str, Path]) -> Workbook:
    """
    EN: Load workbook if exists; otherwise create a new one. Also handles 0-byte/corrupt files.
    中文：加载工作簿；不存在则创建。并处理 0 字节/损坏文件，避免 BadZipFile。
    """
    p = str(path)
    try:
        if os.path.exists(p) and os.path.getsize(p) == 0:
            raise Exception("Empty file (0 bytes)")
        return load_workbook(p)
    except Exception:
        wb = Workbook()
        wb.save(p)
        return wb


def _resolve_target_path(ctx: SkillContext, relative_path: Optional[str], abs_path: Optional[str]) -> Path:
    """
    EN: Prefer abs_path if provided; else resolve relative_path within ctx workspace.
    中文：优先使用 abs_path；否则用 ctx.resolve_path(relative_path)。
    """
    if abs_path:
        return Path(abs_path)
    return ctx.resolve_path(relative_path)


def _ensure_sheet(wb: Workbook, sheet_name: Optional[str]):
    """
    EN: Return named sheet if provided; create if missing. Else return active.
    中文：若指定 sheet_name 则返回该表（不存在则创建）；否则返回 active。
    """
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.create_sheet(title=sheet_name)
    return wb.active


def _parse_cell(cell: str) -> Tuple[int, int]:
    """
    EN: Parse 'A1' -> (row=1, col=1). Minimal parser (no hardcoded sizes).
    中文：解析 'A1' -> (row=1, col=1)。简易解析（无硬编码表大小）。
    """
    # openpyxl provides utility, but keep simple & dependency-free
    cell = cell.strip().upper()
    col_part = ""
    row_part = ""
    for ch in cell:
        if "A" <= ch <= "Z":
            col_part += ch
        elif "0" <= ch <= "9":
            row_part += ch
    if not col_part or not row_part:
        raise ValueError(f"Invalid cell address: {cell}")

    # Convert letters to number: A=1, Z=26, AA=27 ...
    col = 0
    for ch in col_part:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    row = int(row_part)
    if row <= 0 or col <= 0:
        raise ValueError(f"Invalid cell address: {cell}")
    return row, col


def _cell_from_rc(row: int, col: int) -> str:
    """
    EN: (row=1,col=1) -> 'A1'
    中文：(1,1) -> 'A1'
    """
    if row <= 0 or col <= 0:
        raise ValueError("row/col must be >= 1")
    # Convert col number to letters
    letters = ""
    n = col
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(rem + ord("A")) + letters
    return f"{letters}{row}"


def _coerce_2d_rows(data: Any) -> List[List[Any]]:
    """
    EN: Normalize incoming data to 2D rows.
    中文：把输入规范化成二维 rows。
    """
    if data is None:
        return []
    if isinstance(data, list):
        if len(data) == 0:
            return []
        if isinstance(data[0], list):
            return data  # already 2D
        return [data]  # 1D -> single row
    raise ValueError("rows must be a list or list of lists")

# =============================================================================
# excel.read_sheet
# =============================================================================

class ExcelReadSheetInput(PathBindMixin, SkillInput):
    relative_path: str | None = Field(None, description="Relative Excel file path. 相对路径（可空）。")
    sheet_name: Optional[str] = Field(None, description="Sheet name (optional). 表名（可选）。")
    max_rows: int = Field(100, description="Max rows to read. 最大读取行数。")
    max_cols: int = Field(20, description="Max columns to read. 最大读取列数。")

    abs_path: str | None = Field(None, description="Absolute file path. If provided, takes precedence. 绝对路径优先。")

    @model_validator(mode="after")
    def normalize_ext(self):
        if self.relative_path:
            self.relative_path = normalize_file_extension(
                self.relative_path,
                default_ext=".xlsx",
                allowed_exts={".xlsx", ".xls", ".xlsm"},
            )
        return self


class ExcelReadSheetOutput(SkillOutput):
    path: str
    sheet_name: str
    rows: List[List[Any]] = []
    rows_count: int


@register_skill
class ExcelReadSheetSkill(BaseSkill[ExcelReadSheetInput, ExcelReadSheetOutput]):
    spec = SkillSpec(
        name="excel.read_sheet",
        api_name="excel.read",
        aliases=["xlsx.read", "read_excel", "excel.read"],
        description="Read data from an Excel sheet. 读取Excel表格数据。",
        category=SkillCategory.OFFICE,
        input_model=ExcelReadSheetInput,
        output_model=ExcelReadSheetOutput,
        meta=SkillMetadata(
            domain=SkillDomain.OFFICE,
            capabilities={SkillCapability.READ},
            risk_level="normal",
            file_extensions=[".xlsx", ".xls", ".xlsm"],
        ),
        synonyms=[
            "read excel file",
            "read spreadsheet",
            "load excel data",
            "读取Excel",
            "读取表格",
            "加载Excel数据",
        ],
        examples=[
            {"description": "Read Excel sheet", "params": {"relative_path": "data.xlsx"}},
            {"description": "Read specific sheet", "params": {"relative_path": "data.xlsx", "sheet_name": "Sheet1"}},
        ],
        permissions=[SkillPermission(name="file_read", description="Read Excel files")],
        tags=["office", "excel", "read", "表格", "数据", "读取", "Excel"],
    )

    async def run(self, ctx: SkillContext, params: ExcelReadSheetInput) -> ExcelReadSheetOutput:
        target_path = _resolve_target_path(ctx, params.relative_path, params.abs_path)

        if ctx.dry_run:
            return ExcelReadSheetOutput(
                success=True,
                message="[dry_run]",
                path=str(target_path),
                sheet_name=params.sheet_name or "active",
                rows=[],
                rows_count=0,
            )

        try:
            if not target_path.exists():
                return ExcelReadSheetOutput(
                    success=False,
                    message=f"Not found: {target_path}",
                    path=str(target_path),
                    sheet_name="",
                    rows_count=0,
                )

            wb = load_workbook(str(target_path), data_only=True)
            ws = wb[params.sheet_name] if params.sheet_name else wb.active

            rows: List[List[Any]] = []
            for row in ws.iter_rows(
                min_row=1,
                max_row=params.max_rows,
                min_col=1,
                max_col=params.max_cols,
                values_only=True,
            ):
                rows.append(list(row))

            return ExcelReadSheetOutput(
                success=True,
                message=f"Read {len(rows)} rows",
                path=str(target_path),
                sheet_name=ws.title,
                rows=rows,
                rows_count=len(rows),
            )
        except Exception as e:
            return ExcelReadSheetOutput(
                success=False,
                message=str(e),
                path=str(target_path),
                sheet_name="",
                rows_count=0,
            )


# =============================================================================
# excel.write  (SINGLE CELL ONLY)
# =============================================================================

class ExcelWriteCellInput(PathBindMixin, SkillInput):
    relative_path: str | None = Field(None, description="Relative Excel file path (REQUIRED if abs_path not provided). 相对路径（必需，除非提供 abs_path）。")
    sheet_name: Optional[str] = Field(None, description="Sheet name. 表名。")
    cell: str = Field(..., description="Cell address e.g. 'A1'. 单元格地址，例如 A1。")
    value: Any = Field(..., description="Value to write. 写入的值。")

    abs_path: str | None = Field(None, description="Absolute file path. If provided, takes precedence. 绝对路径优先。")

    @model_validator(mode="after")
    def normalize_ext(self):
        if self.relative_path:
            self.relative_path = normalize_file_extension(
                self.relative_path,
                default_ext=".xlsx",
                allowed_exts={".xlsx", ".xls", ".xlsm"},
            )
        return self


class ExcelWriteCellOutput(SkillOutput):
    path: str
    sheet_name: str
    cell: str
    value: Any


@register_skill
class ExcelWriteCellSkill(BaseSkill[ExcelWriteCellInput, ExcelWriteCellOutput]):
    spec = SkillSpec(
        name="excel.write",
        api_name="excel.write",
        aliases=[
            "xlsx.write",
            "excel.write_cell",   # ✅ clear alias to guide LLM & humans
            "write_excel_cell",
        ],
        # IMPORTANT: make boundary explicit to prevent LLM misuse
        description=(
            "Write a SINGLE cell value in an Excel sheet (NOT a whole table). "
            "写入 Excel 的单个单元格（不是写整张表/不是CSV导入）。 "
            "Required params: relative_path (or abs_path), cell, value."
        ),
        category=SkillCategory.OFFICE,
        input_model=ExcelWriteCellInput,
        output_model=ExcelWriteCellOutput,
        meta=SkillMetadata(
            domain=SkillDomain.OFFICE,
            capabilities={SkillCapability.WRITE, SkillCapability.MODIFY},
            risk_level="high",
        ),
        produces_artifact=True,
        artifact_type="document:excel",
        artifact_path_field="path",
        synonyms=[
            "write excel cell",
            "set cell value",
            "update cell",
            "写入单元格",
            "设置单元格值",
            "更新单元格",
        ],
        examples=[
            {
                "description": "Write one cell (A1) in data.xlsx",
                "params": {"relative_path": "data.xlsx", "cell": "A1", "value": "Hello"},
            },
            {
                "description": "Write one cell (B2) in scores.xlsx sheet Sheet1",
                "params": {"relative_path": "scores.xlsx", "sheet_name": "Sheet1", "cell": "B2", "value": 92},
            },
        ],
        permissions=[SkillPermission(name="file_write", description="Write Excel files")],
        tags=["office", "excel", "write", "cell", "表格", "单元格", "写入", "Excel"],
    )

    async def run(self, ctx: SkillContext, params: ExcelWriteCellInput) -> ExcelWriteCellOutput:
        target_path = _resolve_target_path(ctx, params.relative_path, params.abs_path)

        if ctx.dry_run:
            return ExcelWriteCellOutput(
                success=True,
                message="[dry_run]",
                path=str(target_path),
                sheet_name=params.sheet_name or "active",
                cell=params.cell,
                value=params.value,
            )

        try:
            wb = _load_or_create_workbook(target_path)
            ws = _ensure_sheet(wb, params.sheet_name)
            ws[params.cell] = params.value
            wb.save(str(target_path))
            return ExcelWriteCellOutput(
                success=True,
                message=f"Written {params.cell}",
                path=str(target_path),
                sheet_name=ws.title,
                cell=params.cell,
                value=params.value,
            )
        except Exception as e:
            return ExcelWriteCellOutput(
                success=False,
                message=str(e),
                path=str(target_path),
                sheet_name="",
                cell="",
                value="",
            )


# =============================================================================
# excel.append_row
# =============================================================================

class ExcelAppendRowInput(PathBindMixin, SkillInput):
    relative_path: str | None = Field(None, description="Relative Excel file path (REQUIRED if abs_path not provided). 相对路径（必需，除非提供 abs_path）。")
    sheet_name: Optional[str] = Field(None, description="Sheet name. 表名。")
    row_json: Optional[str] = Field(None, description="JSON array string for row values. JSON数组字符串。")

    # Robustness: Allow model to pass 'values', 'data', or 'rows'
    values: Optional[List[Any]] = Field(None, description="Alias for row values. 单行值。")
    data: Optional[List[Any]] = Field(None, description="Alias for row values. 单行值。")
    rows: Optional[List[Any]] = Field(None, description="Alias for row values (support batch). 多行/批量。")

    abs_path: str | None = Field(None, description="Absolute file path. If provided, takes precedence. 绝对路径优先。")

    @model_validator(mode="before")
    def handle_dict_inputs(cls, values):
        """
        EN: If LLM passes a dict (SkillOutput) into rows/values/data, try to extract usable payload.
        中文：若 LLM 把上游 SkillOutput(dict) 塞进 rows/values/data，尝试提取可用数据。
        """
        if not isinstance(values, dict):
            return values

        for key in ["rows", "values", "data"]:
            if key in values and isinstance(values[key], dict):
                incoming_dict = values[key]

                if "value" in incoming_dict:
                    val = incoming_dict["value"]
                    values[key] = [[val]] if key == "rows" else [val]
                    return values

                if "rows" in incoming_dict and isinstance(incoming_dict["rows"], list):
                    values[key] = incoming_dict["rows"]
                    return values

        return values

    @model_validator(mode="after")
    def normalize_ext(self):
        if self.relative_path:
            self.relative_path = normalize_file_extension(
                self.relative_path,
                default_ext=".xlsx",
                allowed_exts={".xlsx", ".xls", ".xlsm"},
            )
        return self


class ExcelAppendRowOutput(SkillOutput):
    path: str
    sheet_name: str
    row: List[Any]


@register_skill
class ExcelAppendRowSkill(BaseSkill[ExcelAppendRowInput, ExcelAppendRowOutput]):
    spec = SkillSpec(
        name="excel.append",
        api_name="excel.append",
        aliases=["xlsx.append", "append_excel", "excel.append_row"],
        description="Append a row (or rows) to an Excel sheet. Required: relative_path or abs_path. 向Excel表格追加行数据。必需提供文件路径。",
        category=SkillCategory.OFFICE,
        input_model=ExcelAppendRowInput,
        output_model=ExcelAppendRowOutput,
        meta=SkillMetadata(
            domain=SkillDomain.OFFICE,
            capabilities={SkillCapability.WRITE, SkillCapability.MODIFY},
            risk_level="high",
        ),
        produces_artifact=True,
        artifact_type="document:excel",
        artifact_path_field="path",
        synonyms=[
            "append excel row",
            "add row to excel",
            "insert row",
            "追加行",
            "添加行",
            "插入行",
        ],
        examples=[
            {"description": "Append single row", "params": {"relative_path": "data.xlsx", "values": ["Name", "Age", "City"]}},
            {"description": "Append multiple rows", "params": {"relative_path": "data.xlsx", "rows": [["Alice", 30], ["Bob", 25]]}},
        ],
        permissions=[SkillPermission(name="file_write", description="Write Excel files")],
        tags=["office", "excel", "append", "表格", "追加", "行", "Excel"],
    )

    async def run(self, ctx: SkillContext, params: ExcelAppendRowInput) -> ExcelAppendRowOutput:
        target_path = _resolve_target_path(ctx, params.relative_path, params.abs_path)

        # Resolve row data (robust)
        try:
            raw_data: Any = None
            if params.values is not None:
                raw_data = params.values
            elif params.data is not None:
                raw_data = params.data
            elif params.rows is not None:
                raw_data = params.rows
            elif params.row_json is not None:
                raw_data = params.row_json if isinstance(params.row_json, list) else json.loads(params.row_json)
            else:
                raise ValueError("Missing row data. Provide 'row_json', 'values', 'data', or 'rows'.")

            rows_to_append = _coerce_2d_rows(raw_data)
        except Exception as e:
            return ExcelAppendRowOutput(success=False, message=f"Invalid data format: {str(e)}", path="", sheet_name="", row=[])

        if ctx.dry_run:
            return ExcelAppendRowOutput(
                success=True,
                message="[dry_run]",
                path=str(target_path),
                sheet_name=params.sheet_name or "active",
                row=rows_to_append[0] if rows_to_append else [],
            )

        try:
            wb = _load_or_create_workbook(target_path)
            ws = _ensure_sheet(wb, params.sheet_name)
            for r in rows_to_append:
                ws.append(r)
            wb.save(str(target_path))

            return ExcelAppendRowOutput(
                success=True,
                message=f"Appended {len(rows_to_append)} row(s)",
                path=str(target_path),
                sheet_name=ws.title,
                row=rows_to_append[0] if rows_to_append else [],
            )
        except Exception as e:
            return ExcelAppendRowOutput(success=False, message=str(e), path=str(target_path), sheet_name="", row=[])


# =============================================================================
# NEW: excel.write_table  (WRITE A WHOLE TABLE / BATCH CELLS)
# =============================================================================

class ExcelWriteTableInput(PathBindMixin, SkillInput):
    relative_path: str | None = Field(None, description="Relative Excel file path (REQUIRED if abs_path not provided). 相对路径（必需，除非提供 abs_path）。")
    sheet_name: Optional[str] = Field(None, description="Sheet name. 表名（可选）。")

    # EN: 2D rows data
    # 中文：二维表数据
    rows: Optional[List[List[Any]]] = Field(None, description="2D table rows. 二维表数据。")

    # EN: Where to start writing (default A1)
    # 中文：从哪个单元格开始写（默认 A1）
    start_cell: str = Field("A1", description="Start cell e.g. A1. 起始单元格。")

    # If true, clear sheet before writing
    clear_sheet: bool = Field(False, description="Clear the sheet before writing. 写入前清空工作表。")

    abs_path: str | None = Field(None, description="Absolute file path. If provided, takes precedence. 绝对路径优先。")

    @model_validator(mode="after")
    def normalize_ext(self):
        if self.relative_path:
            self.relative_path = normalize_file_extension(
                self.relative_path,
                default_ext=".xlsx",
                allowed_exts={".xlsx", ".xls", ".xlsm"},
            )
        return self

    @model_validator(mode="after")
    def validate_payload(self):
        if self.rows is None:
            raise ValueError("Missing 'rows'. 必须提供 rows。")

        if not isinstance(self.rows, list):
            raise ValueError("'rows' must be a list.")

        if len(self.rows) == 0:
            raise ValueError("'rows' is empty.")

        # rows = [1,2,3]  → 合法（单行）
        # rows = [[1,2],[3,4]] → 合法
        # rows = [[]] / [[],[]] → 非法
        if isinstance(self.rows[0], list):
            if all(len(r) == 0 for r in self.rows):
                raise ValueError("'rows' contains no data (all rows empty).")
        else:
            # 1D row is acceptable
            pass

        return self


class ExcelWriteTableOutput(SkillOutput):
    path: str
    sheet_name: str
    start_cell: str
    rows_written: int
    cols_written: int


@register_skill
class ExcelWriteTableSkill(BaseSkill[ExcelWriteTableInput, ExcelWriteTableOutput]):
    spec = SkillSpec(
        name="excel.write_table",
        api_name="excel.write_table",
        aliases=["excel.from_csv", "xlsx.write_table", "write_excel_table", "excel.create_with_data"],
        description=(
            "Write a whole TABLE into Excel starting at a cell (A1 by default). "
            "RECOMMENDED for creating Excel files with header and multiple rows in ONE operation. "
            "Supports rows (2D array). Required: relative_path or abs_path. "
            "写入整张表到Excel（默认从A1开始）。推荐用于一次性写入表头+多行数据。"
            "支持 rows(二维数组)。必需提供文件路径。"
        ),
        category=SkillCategory.OFFICE,
        input_model=ExcelWriteTableInput,
        output_model=ExcelWriteTableOutput,
        meta=SkillMetadata(
            domain=SkillDomain.OFFICE,
            capabilities={SkillCapability.WRITE, SkillCapability.MODIFY},
            risk_level="high",
            file_extensions=[".xlsx", ".xlsm", ".xls"],
            priority=70,  # 提升优先级，引导 LLM 优先选择
        ),
        produces_artifact=True,
        artifact_type="document:excel",
        artifact_path_field="path",
        synonyms=[
            "create excel with data",
            "write excel table",
            "import csv to excel",
            "create excel with header and rows",
            "batch write excel",
            "写表",
            "写入表格",
            "把CSV写入Excel",
            "创建Excel并写入数据",
            "创建带表头的Excel",
            "批量写入Excel",
        ],
        examples=[
            {
                "description": "Create Excel with header and 3 data rows",
                "params": {
                    "relative_path": "sales.xlsx",
                    "rows": [
                        ["Date", "Product", "Quantity", "Price"],
                        ["2025-01-01", "Apple", 10, 3.5],
                        ["2025-01-02", "Banana", 8, 2.0],
                        ["2025-01-03", "Apple", 6, 3.5]
                    ],
                    "start_cell": "A1",
                },
            },
            {
                "description": "Write table rows to scores.xlsx",
                "params": {
                    "relative_path": "scores.xlsx",
                    "rows": [["name", "score"], ["Alice", 85], ["Bob", 92], ["Charlie", 78]],
                    "start_cell": "A1",
                },
            },
        ],
        permissions=[SkillPermission(name="file_write", description="Write Excel files")],
        tags=["office", "excel", "table", "csv", "batch", "表格", "写入", "批量", "Excel"],
    )

    async def run(self, ctx: SkillContext, params: ExcelWriteTableInput) -> ExcelWriteTableOutput:
        target_path = _resolve_target_path(ctx, params.relative_path, params.abs_path)

        # ---------- Pre-validate / normalize ----------
        try:
            rows_data = _coerce_2d_rows(params.rows)
            if not rows_data or all((r is None or len(r) == 0) for r in rows_data):
                return ExcelWriteTableOutput(
                    success=False,
                    message="No valid rows provided (empty table).",
                    path=str(target_path),
                    sheet_name="",
                    start_cell=params.start_cell,
                    rows_written=0,
                    cols_written=0,
                )
        except Exception as e:
            return ExcelWriteTableOutput(
                success=False,
                message=f"Invalid rows: {str(e)}",
                path=str(target_path),
                sheet_name="",
                start_cell=params.start_cell,
                rows_written=0,
                cols_written=0,
            )

        r_count = len(rows_data)
        c_count = max((len(r) for r in rows_data if r is not None), default=0)

        if ctx.dry_run:
            return ExcelWriteTableOutput(
                success=True,
                message=f"[dry_run] Would write table {r_count}x{c_count} at {params.start_cell}",
                path=str(target_path),
                sheet_name=params.sheet_name or "active",
                start_cell=params.start_cell,
                rows_written=r_count,
                cols_written=c_count,
            )

        try:
            # ---------- Execute ----------
            wb = _load_or_create_workbook(target_path)
            ws = _ensure_sheet(wb, params.sheet_name)

            # Optional clear (simple & safe)
            if params.clear_sheet:
                if ws.max_row and ws.max_row > 0:
                    ws.delete_rows(1, ws.max_row)

            start_r, start_c = _parse_cell(params.start_cell)

            max_cols = 0
            for i, row in enumerate(rows_data):
                row = row or []
                max_cols = max(max_cols, len(row))
                for j, val in enumerate(row):
                    ws.cell(row=start_r + i, column=start_c + j, value=val)

            # Ensure parent dir exists (file.py style)
            Path(target_path).parent.mkdir(parents=True, exist_ok=True)

            wb.save(str(target_path))

            # ---------- Post-verify ----------
            if not Path(target_path).exists():
                return ExcelWriteTableOutput(
                    success=False,
                    message=f"Validator Error: File not found at {target_path} after save.",
                    path=str(target_path),
                    sheet_name="",
                    start_cell=params.start_cell,
                    rows_written=0,
                    cols_written=0,
                )

            # Verify a sentinel cell (top-left of written range)
            # NOTE: This is a lightweight integrity check, avoids re-reading whole table.
            try:
                verify_wb = load_workbook(str(target_path), data_only=False)
                verify_ws = verify_wb[ws.title] if ws.title in verify_wb.sheetnames else verify_wb.active
                expected = (rows_data[0][0] if rows_data and rows_data[0] else None)
                actual = verify_ws.cell(row=start_r, column=start_c).value
                if actual != expected:
                    return ExcelWriteTableOutput(
                        success=False,
                        message=f"Validator Error: Cell {params.start_cell} mismatch after save (actual={actual!r}, expected={expected!r}).",
                        path=str(target_path),
                        sheet_name=verify_ws.title,
                        start_cell=params.start_cell,
                        rows_written=0,
                        cols_written=0,
                    )
            except Exception:
                # Don't fail hard on verification read errors; main write succeeded.
                pass

            return ExcelWriteTableOutput(
                success=True,
                message=f"Wrote table {r_count}x{max_cols} starting at {params.start_cell}",
                path=str(target_path),
                sheet_name=ws.title,
                start_cell=params.start_cell,
                rows_written=r_count,
                cols_written=max_cols,
                # 可选：如果你 SkillOutput 支持 FS 元数据字段，也可以在这里加
                # fs_operation="modified" if Path(target_path).exists() else "created",
                # fs_path=params.relative_path or os.path.basename(params.abs_path),
                # fs_type="file",
            )

        except Exception as e:
            return ExcelWriteTableOutput(
                success=False,
                message=str(e),
                path=str(target_path),
                sheet_name="",
                start_cell=params.start_cell,
                rows_written=0,
                cols_written=0,
            )


# =============================================================================
# NEW: excel.add_formula_column (添加计算列)
# =============================================================================

class ExcelAddFormulaColumnInput(PathBindMixin, SkillInput):
    relative_path: str | None = Field(None, description="Relative Excel file path (REQUIRED if abs_path not provided). 相对路径（必需，除非提供 abs_path）。")
    sheet_name: Optional[str] = Field(None, description="Sheet name. 表名（可选）。")
    
    column: Optional[str] = Field(None, description="Column letter (e.g. 'E' for column E). Auto-infer from formula if not provided. 列字母（可选，未提供时自动推断）。")
    header: Optional[str] = Field(None, description="Header name for the new column. 新列的表头名称。")
    formula_template: str = Field(
        ..., 
        description=(
            "**REQUIRED** Formula template with {row} placeholder. "
            "Example: '=C{row}*D{row}' to multiply columns C and D. "
            "必需参数。公式模板，使用 {row} 占位符，例如 =C{row}*D{row}。"
        )
    )
    
    start_row: int = Field(2, description="First data row (usually 2 if row 1 is header). 起始数据行（通常为2，如果第1行是表头）。")
    end_row: Optional[int] = Field(None, description="Last data row (auto-detect if None). 结束行（None则自动检测）。")
    
    abs_path: str | None = Field(None, description="Absolute file path. If provided, takes precedence. 绝对路径优先。")

    @model_validator(mode="before")
    def map_formula_aliases(cls, values):
        """
        同义字段自动映射（Auto-repair 策略：轻量纠错，不做业务推断）
        
        映射规则：
        1. formula_template 别名：
           formula/template/expression/calc/calculation → formula_template
        
        2. column 别名：
           col/target_column/result_column/output_column/dest_column/to_column → column
        
        覆盖 LLM 常见的参数命名变体，提高容错性。
        """
        if not isinstance(values, dict):
            return values
        
        # 1. formula_template 同义字段映射
        if "formula_template" not in values:
            for alias in ["formula", "template", "expression", "calc", "calculation"]:
                if alias in values:
                    values["formula_template"] = values[alias]
                    logger.debug(f"{cls.__name__}: Mapped alias '{alias}' → 'formula_template'")
                    break
        
        # 2. column 同义字段映射
        if "column" not in values:
            for alias in ["col", "target_column", "result_column", "output_column", "dest_column", "to_column"]:
                if alias in values:
                    values["column"] = values[alias]
                    logger.debug(f"{cls.__name__}: Mapped alias '{alias}' → 'column'")
                    break
        
        return values
    
    @model_validator(mode="after")
    def auto_repair_formula(self):
        """
        公式自动修复（轻量纠错）
        
        修复规则：
        1. 不以 = 开头 → 自动补 =
        2. 没有 {row} 占位符 → 尝试把具体行号（如 C2*D2）改为 {row}
        """
        import re
        
        formula = self.formula_template
        
        # 修复1: 补 =
        if not formula.startswith("="):
            formula = "=" + formula
            logger.debug(f"{self.__class__.__name__}: Auto-repaired: Added '=' prefix")
        
        # 修复2: 替换具体行号为 {row}
        # 检测模式：字母+数字（如 A1, C2, AA10）
        if "{row}" not in formula:
            # 匹配 Excel 单元格引用（如 C2, D2, AA10）
            pattern = r'([A-Z]+)(\d+)'
            matches = re.findall(pattern, formula)
            
            if matches:
                # 检查是否所有引用的行号一致（如都是2）
                row_numbers = [int(m[1]) for m in matches]
                if len(set(row_numbers)) == 1:
                    # 所有引用同一行，替换为 {row}
                    formula = re.sub(pattern, r'\1{row}', formula)
                    logger.debug(
                        f"{self.__class__.__name__}: Auto-repaired: "
                        f"Replaced row number {row_numbers[0]} with {{row}}"
                    )
        
        self.formula_template = formula
        
        # 扩展名规范化
        if self.relative_path:
            self.relative_path = normalize_file_extension(
                self.relative_path,
                default_ext=".xlsx",
                allowed_exts={".xlsx", ".xls", ".xlsm"},
            )
        
        return self


class ExcelAddFormulaColumnOutput(SkillOutput):
    path: str
    sheet_name: str
    column: str
    rows_filled: int


@register_skill
class ExcelAddFormulaColumnSkill(BaseSkill[ExcelAddFormulaColumnInput, ExcelAddFormulaColumnOutput]):
    @staticmethod
    def _infer_column(ws, params: ExcelAddFormulaColumnInput) -> str:
        """
        智能推断目标列（方案6：从 formula_template 推断）
        
        推断逻辑：
        1. 护栏A：如果有 header，尝试匹配现有表头位置
        2. 如果没有 header 或匹配失败，从 formula_template 提取引用列，取 max+1
        3. 护栏B：检查已用列范围，避免覆盖已有列
        """
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        # === 护栏A：优先用 header 匹配现有表头 ===
        if params.header and ws.max_row and ws.max_row >= 1:
            # 读取第一行所有表头
            for col_idx in range(1, ws.max_column + 1 if ws.max_column else 100):
                cell_value = ws.cell(1, col_idx).value
                if cell_value and str(cell_value).strip().lower() == params.header.strip().lower():
                    # 找到匹配的表头，直接返回该列
                    matched_col = get_column_letter(col_idx)
                    logger.debug(f"ExcelAddFormulaColumnSkill: Matched header '{params.header}' at column {matched_col}")
                    return matched_col
        
        # === 从 formula_template 提取引用列 ===
        # 匹配 Excel 单元格引用（如 C{row}, D2, AA{row}）
        pattern = r'([A-Z]+)(?:\{row\}|\d+)'
        matches = re.findall(pattern, params.formula_template)
        
        max_col_idx = 0
        if matches:
            # 找到公式中引用的最大列
            for col_letter in matches:
                try:
                    col_idx = column_index_from_string(col_letter)
                    max_col_idx = max(max_col_idx, col_idx)
                except:
                    pass
        
        # === 护栏B：检查已用列，避免冲突 ===
        # 候选列 = max(公式引用列) + 1
        candidate_col_idx = max_col_idx + 1 if max_col_idx > 0 else 5  # 默认 E 列
        
        # 如果候选列 <= 当前 sheet 最大列，说明可能被占用，使用 sheet 最大列 + 1
        if ws.max_column and candidate_col_idx <= ws.max_column:
            candidate_col_idx = ws.max_column + 1
            logger.debug(f"ExcelAddFormulaColumnSkill: Candidate column occupied, using max_column + 1")
        
        inferred_col = get_column_letter(candidate_col_idx)
        return inferred_col
    
    spec = SkillSpec(
        name="excel.add_formula_column",
        api_name="excel.add_formula_column",
        aliases=["excel.fill_formula", "excel.add_calculated_column", "excel.compute_column"],
        description=(
            "Add a calculated column with formula to Excel. "
            "Auto-fills formula to all data rows. "
            "**MUST provide formula_template with {row} placeholder** (e.g. '=C{row}*D{row}'). "
            "添加计算列到Excel，自动填充公式到所有数据行。"
            "**必需提供 formula_template，使用 {row} 占位符**，例如 =C{row}*D{row}。"
        ),
        category=SkillCategory.OFFICE,
        input_model=ExcelAddFormulaColumnInput,
        output_model=ExcelAddFormulaColumnOutput,
        meta=SkillMetadata(
            domain=SkillDomain.OFFICE,
            capabilities={SkillCapability.WRITE, SkillCapability.MODIFY},
            risk_level="high",
            file_extensions=[".xlsx", ".xlsm"],
            priority=60,
        ),
        produces_artifact=True,
        artifact_type="document:excel",
        artifact_path_field="path",
        synonyms=[
            "add calculated column",
            "fill down formula",
            "add formula column",
            "compute column",
            "添加计算列",
            "填充公式",
            "添加公式列",
            "计算列",
        ],
        examples=[
            {
                "description": "Multiply columns: Total = Quantity (C) × Price (D)",
                "params": {
                    "relative_path": "sales.xlsx",
                    "column": "E",
                    "header": "Total",
                    "formula_template": "=C{row}*D{row}",  # ← REQUIRED with {row}
                    "start_row": 2
                },
            },
            {
                "description": "Calculate percentage: Tax = Total (E) × 10%",
                "params": {
                    "relative_path": "sales.xlsx",
                    "column": "F",
                    "header": "Tax",
                    "formula_template": "=E{row}*0.1",  # ← REQUIRED with {row}
                    "start_row": 2
                },
            },
            {
                "description": "Add columns: Total = Price (C) + Tax (D)",
                "params": {
                    "relative_path": "report.xlsx",
                    "column": "E",
                    "formula_template": "=C{row}+D{row}",  # ← Can omit '=' (auto-added)
                    "start_row": 2
                },
            },
        ],
        permissions=[SkillPermission(name="file_write", description="Write Excel files")],
        tags=["office", "excel", "formula", "calculate", "表格", "公式", "计算", "Excel"],
    )

    async def run(self, ctx: SkillContext, params: ExcelAddFormulaColumnInput) -> ExcelAddFormulaColumnOutput:
        target_path = _resolve_target_path(ctx, params.relative_path, params.abs_path)

        # dry_run 只展示，不要硬塞默认列
        if ctx.dry_run:
            return ExcelAddFormulaColumnOutput(
                success=True,
                message="[dry_run]",
                path=str(target_path),
                sheet_name=params.sheet_name or "active",
                column=params.column,
                rows_filled=0,
            )

        try:
            if not target_path.exists():
                return ExcelAddFormulaColumnOutput(
                    success=False,
                    message=f"File not found: {target_path}",
                    path=str(target_path),
                    sheet_name="",
                    column=params.column,
                    rows_filled=0,
                )

            wb = load_workbook(str(target_path))
            ws = _ensure_sheet(wb, params.sheet_name)

            # ✅ 不要修改 params，使用局部变量更稳
            column = params.column or self._infer_column(ws, params)
            if not column:
                raise ValueError("column could not be inferred")

            # 1) header
            if params.header:
                ws[f"{column}1"] = params.header

            # 2) end_row
            end_row = params.end_row
            if end_row is None:
                end_row = ws.max_row
                if not end_row or end_row < params.start_row:
                    return ExcelAddFormulaColumnOutput(
                        success=False,
                        message="No data rows found in sheet",
                        path=str(target_path),
                        sheet_name=ws.title,
                        column=column,
                        rows_filled=0,
                    )

            # 3) fill formula
            rows_filled = 0
            for row in range(params.start_row, end_row + 1):
                formula = params.formula_template.replace("{row}", str(row))
                ws[f"{column}{row}"] = formula
                rows_filled += 1

            wb.save(str(target_path))

            return ExcelAddFormulaColumnOutput(
                success=True,
                message=f"Added formula column {column} with {rows_filled} rows",
                path=str(target_path),
                sheet_name=ws.title,
                column=column,
                rows_filled=rows_filled,
            )

        except Exception as e:
            return ExcelAddFormulaColumnOutput(
                success=False,
                message=str(e),
                path=str(target_path),
                sheet_name="",
                column=params.column,
                rows_filled=0,
            )


# =============================================================================
# excel.add_chart
# =============================================================================

class ExcelAddChartInput(PathBindMixin, SkillInput):
    relative_path: str | None = Field(None, description="Relative Excel file path. 相对路径。")
    abs_path: str | None = Field(None, description="Absolute path (overrides relative_path).")
    sheet_name: Optional[str] = Field(None, description="Source data sheet name. 数据所在表名。")
    chart_type: str = Field("bar", description="Chart type: bar, line, pie, scatter, area. 图表类型。")
    title: Optional[str] = Field(None, description="Chart title. 图表标题。")
    data_range: Optional[str] = Field(
        None,
        description="Data range in A1 notation, e.g. 'A1:D10'. Auto-detected if omitted. 数据范围。",
    )
    x_column: Optional[str] = Field(
        None,
        description="Column letter for X-axis / categories, e.g. 'A'. 分类轴列。",
    )
    y_columns: Optional[str] = Field(
        None,
        description="Comma-separated column letters for Y-axis series, e.g. 'B,C'. 数据列。",
    )
    target_cell: str = Field("E2", description="Cell where the chart top-left corner is placed. 图表放置位置。")
    chart_sheet: Optional[str] = Field(
        None,
        description="If set, place chart on this sheet (created if missing). 图表放置的表名。",
    )
    width: float = Field(15, description="Chart width in cm.")
    height: float = Field(10, description="Chart height in cm.")

    @model_validator(mode="before")
    @classmethod
    def normalize_ext(cls, values):
        return normalize_file_extension(values, target_ext=".xlsx")


class ExcelAddChartOutput(SkillOutput):
    path: str = ""
    sheet_name: str = ""
    chart_type: str = ""
    series_count: int = 0


@register_skill
class ExcelAddChartSkill(BaseSkill[ExcelAddChartInput, ExcelAddChartOutput]):
    spec = SkillSpec(
        name="excel.add_chart",
        api_name="excel.add_chart",
        aliases=["excel.chart", "excel.create_chart"],
        description="Add a chart (bar, line, pie, scatter, area) to an Excel file. 在 Excel 中添加图表。",
        category=SkillCategory.OFFICE,
        input_model=ExcelAddChartInput,
        output_model=ExcelAddChartOutput,
        meta=SkillMetadata(
            domain=SkillDomain.OFFICE,
            capabilities={SkillCapability.MODIFY, SkillCapability.CREATE},
            risk_level="normal",
        ),
        produces_artifact=True,
        artifact_type="document:excel",
        artifact_path_field="path",
        synonyms=[
            "create chart",
            "add graph",
            "make chart",
            "生成图表",
            "添加图表",
            "创建图表",
            "柱状图",
            "折线图",
            "饼图",
            "数据可视化",
        ],
        examples=[
            {
                "description": "Add bar chart",
                "params": {"relative_path": "report.xlsx", "chart_type": "bar", "title": "Sales"},
            },
            {
                "description": "Add pie chart on specific range",
                "params": {
                    "relative_path": "data.xlsx",
                    "chart_type": "pie",
                    "data_range": "A1:B5",
                    "title": "Distribution",
                },
            },
        ],
        tags=["excel", "chart", "graph", "图表", "可视化", "报表"],
    )

    async def run(self, ctx: SkillContext, params: ExcelAddChartInput) -> ExcelAddChartOutput:
        from openpyxl.chart import (
            BarChart,
            LineChart,
            PieChart,
            ScatterChart,
            AreaChart,
            Reference,
        )
        from openpyxl.utils import get_column_letter, column_index_from_string

        target_path = _resolve_target_path(ctx, params.relative_path, params.abs_path)

        try:
            wb = _load_or_create_workbook(target_path)
            ws = _ensure_sheet(wb, params.sheet_name)

            # Auto-detect data range if not specified
            min_row = ws.min_row or 1
            max_row = ws.max_row or 1
            min_col = ws.min_column or 1
            max_col = ws.max_column or 1

            if params.data_range:
                # Parse A1:D10 style range
                parts = params.data_range.replace(" ", "").split(":")
                if len(parts) == 2:
                    r1, c1 = _parse_cell(parts[0])
                    r2, c2 = _parse_cell(parts[1])
                    min_row, min_col = r1, c1
                    max_row, max_col = r2, c2

            # Determine category (X) and data (Y) columns
            if params.x_column:
                cat_col = column_index_from_string(params.x_column.upper())
            else:
                cat_col = min_col  # First column = categories

            if params.y_columns:
                y_cols = [
                    column_index_from_string(c.strip().upper())
                    for c in params.y_columns.split(",")
                ]
            else:
                # All columns except the category column
                y_cols = [c for c in range(min_col, max_col + 1) if c != cat_col]

            if not y_cols:
                return ExcelAddChartOutput(
                    success=False,
                    message="No data columns found for chart series.",
                    path=str(target_path),
                    sheet_name=ws.title,
                    chart_type=params.chart_type,
                )

            # Create chart object
            chart_map = {
                "bar": BarChart,
                "column": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "scatter": ScatterChart,
                "area": AreaChart,
            }
            chart_cls = chart_map.get(params.chart_type.lower(), BarChart)
            chart = chart_cls()

            if params.title:
                chart.title = params.title
            chart.width = params.width
            chart.height = params.height

            # Data starts from row after header (assume row 1 = header)
            data_min_row = min_row + 1 if max_row > min_row else min_row

            # Categories reference (X-axis labels)
            cats = Reference(ws, min_col=cat_col, min_row=data_min_row, max_row=max_row)

            # Add each Y column as a series
            series_count = 0
            for yc in y_cols:
                data_ref = Reference(ws, min_col=yc, min_row=min_row, max_row=max_row)
                chart.add_data(data_ref, titles_from_data=True)
                series_count += 1

            if not isinstance(chart, PieChart):
                chart.set_categories(cats)

            # Place chart on target sheet
            if params.chart_sheet:
                chart_ws = _ensure_sheet(wb, params.chart_sheet)
            else:
                chart_ws = ws

            chart_ws.add_chart(chart, params.target_cell)

            wb.save(str(target_path))

            return ExcelAddChartOutput(
                success=True,
                message=f"Added {params.chart_type} chart with {series_count} series at {params.target_cell}",
                path=str(target_path),
                sheet_name=chart_ws.title,
                chart_type=params.chart_type,
                series_count=series_count,
            )

        except Exception as e:
            return ExcelAddChartOutput(
                success=False,
                message=str(e),
                path=str(target_path),
                sheet_name=params.sheet_name or "",
                chart_type=params.chart_type,
            )
